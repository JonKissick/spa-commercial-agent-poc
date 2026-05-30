import sys
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.calculators.schemas import NpvCalculationRequest, ScenarioAssumptions
from app.main import app
from app.workbooks.npv_workbook import build_npv_workbook


def _request() -> NpvCalculationRequest:
    return NpvCalculationRequest(
        contract_role="buyer",
        delivery_basis="fob",
        discount_rate=0.1,
        contract_years=2,
        currency="USD",
        unit="MMBtu",
        scenarios=[
            ScenarioAssumptions(
                scenario_name="base",
                annual_volume=100,
                contract_price=8,
                market_price=12,
                freight_cost=1,
                boil_off_cost=0.2,
                port_canal_cost=0.1,
                regas_terminal_cost=0.4,
                downstream_cost=0.3,
                annual_fixed_costs=10,
            )
        ],
    )


def test_build_npv_workbook_contains_expected_sheets() -> None:
    from app.calculators.npv import calculate_npv

    workbook_bytes = build_npv_workbook(calculate_npv(_request()))
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assert workbook.sheetnames == ["Summary", "Scenario Results", "Sensitivities", "Break-even", "Warnings and Limits"]
    assert workbook["Summary"]["A1"].value == "Field"
    assert workbook["Scenario Results"]["A2"].value == "base"
    assert workbook["Scenario Results"]["D2"].value == 329.75


def test_npv_workbook_endpoint_returns_xlsx() -> None:
    client = TestClient(app)

    response = client.post("/calculators/npv/workbook", json=_request().model_dump(mode="json"))

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert "Scenario Results" in workbook.sheetnames
