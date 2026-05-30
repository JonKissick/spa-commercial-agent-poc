from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.calculators.schemas import NpvCalculationResponse


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


def build_npv_workbook(response: NpvCalculationResponse) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_summary(summary, response)
    _write_scenarios(workbook.create_sheet("Scenario Results"), response)
    _write_sensitivities(workbook.create_sheet("Sensitivities"), response)
    _write_break_even(workbook.create_sheet("Break-even"), response)
    _write_notes(workbook.create_sheet("Warnings and Limits"), response)

    for worksheet in workbook.worksheets:
        _autosize_columns(worksheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_summary(sheet, response: NpvCalculationResponse) -> None:
    rows = [
        ("Status", response.calculation_status.value),
        ("Contract role", response.contract_role.value),
        ("Delivery basis", response.delivery_basis.value),
        ("Currency", response.currency),
        ("Unit", response.unit),
        ("Discount rate", response.discount_rate),
        ("Contract years", response.contract_years),
        ("Workbook status", "Draft local POC output; requires analyst validation."),
    ]
    sheet.append(["Field", "Value"])
    _style_header(sheet, 1)
    for row in rows:
        sheet.append(row)


def _write_scenarios(sheet, response: NpvCalculationResponse) -> None:
    sheet.append(
        [
            "Scenario",
            "Annual unit margin",
            "Annual cash flow",
            "NPV",
            "Undiscounted total cash flow",
            "Formula",
            "Included costs",
            "Excluded costs",
        ]
    )
    _style_header(sheet, 1)
    for result in response.scenario_results:
        sheet.append(
            [
                result.scenario_name.value,
                result.annual_unit_margin,
                result.annual_cash_flow,
                result.npv,
                result.undiscounted_total_cash_flow,
                result.formula_used,
                ", ".join(result.included_costs),
                ", ".join(result.excluded_costs),
            ]
        )


def _write_sensitivities(sheet, response: NpvCalculationResponse) -> None:
    sheet.append(["Scenario", "Variable", "Shift", "Resulting NPV", "Resulting annual unit margin", "Note"])
    _style_header(sheet, 1)
    for table in response.sensitivity_tables:
        for point in table.points:
            sheet.append(
                [
                    point.scenario_name.value,
                    point.variable,
                    point.shift,
                    point.resulting_npv,
                    point.resulting_annual_unit_margin,
                    point.note or "",
                ]
            )


def _write_break_even(sheet, response: NpvCalculationResponse) -> None:
    sheet.append(
        [
            "Scenario",
            "Break-even market price",
            "Break-even contract price",
            "Break-even freight cost",
            "Break-even annual volume",
            "Notes",
            "Warnings",
        ]
    )
    _style_header(sheet, 1)
    for result in response.break_even_results:
        sheet.append(
            [
                result.scenario_name.value,
                result.break_even_market_price,
                result.break_even_contract_price,
                result.break_even_freight_cost,
                result.break_even_annual_volume,
                " | ".join(result.notes),
                " | ".join(result.warnings),
            ]
        )


def _write_notes(sheet, response: NpvCalculationResponse) -> None:
    sheet.append(["Type", "Message"])
    _style_header(sheet, 1)
    for warning in response.warnings:
        sheet.append(["Warning", warning])
        for cell in sheet[sheet.max_row]:
            cell.fill = WARNING_FILL
    for limitation in response.limitations:
        sheet.append(["Limitation", limitation])


def _style_header(sheet, row_number: int) -> None:
    for cell in sheet[row_number]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    sheet.freeze_panes = "A2"


def _autosize_columns(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
